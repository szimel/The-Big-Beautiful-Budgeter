import csv
import json
import argparse
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


MIN_SUPPORTED_DATE = date(2026, 6, 1)
CENT = Decimal("0.01")


def parse_csv_date(date_value: str) -> datetime:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime((date_value or "").strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {date_value}")


def format_short_date(dt: datetime) -> str:
    return f"{dt.month}/{dt.day}/{str(dt.year)[2:]}"


def normalize_cardholder(name: str) -> str:
    text = (name or "").strip()
    lower = text.casefold()
    if lower.startswith("samuel"):
        return "Samuel"
    if lower.startswith("kamrie"):
        return "Kamrie"
    if not text:
        return ""
    return text.split()[0].title()


def parse_amount(value: str) -> float:
    raw = (value or "").replace(",", "").replace("$", "").strip()
    return float(raw) if raw else 0.0


def amount_to_cents(value: Any) -> int | None:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None

    if not amount.is_finite() or amount <= 0:
        return None

    cents = amount * 100
    if cents != cents.to_integral_value():
        return None
    return int(cents)


def cents_to_amount(cents: int) -> float:
    return float((Decimal(cents) / 100).quantize(CENT))


def format_cents(cents: int) -> str:
    return f"{cents_to_amount(cents):.2f}"


def parse_points(value: str) -> int:
    raw = (value or "").replace(",", "").strip()
    return int(float(raw)) if raw else 0


def normalize_time(time_value: str) -> str:
    return (time_value or "").strip().upper()


def make_transaction_id(date_str: str, time_str: str) -> str:
    return f"{date_str}|{time_str}"


def parse_transaction_datetime(transaction: dict) -> datetime:
    dt = datetime.strptime(transaction["Date"], "%m/%d/%y")
    time_value = transaction.get("Time", "").strip().upper()
    if time_value:
        try:
            t = datetime.strptime(time_value, "%I:%M %p")
            dt = dt.replace(hour=t.hour, minute=t.minute)
        except ValueError:
            pass
    return dt


def load_json(path: Path, fallback: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return fallback
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)


def ensure_categories_file(path: Path) -> dict[str, float]:
    default_payload = {
        "category_budgets": {
            "Electricity": 100.00,
            "Tooth Payment": 400.00,
            "Rent": 2065.00,
            "Student Loans": 800.00,
            "Insurance": 293.66,
            "Emergency Fund": 500.00,
            "Grocery": 400.00,
            "Gas": 200.00,
            "S \"Fun Money\"": 200.00,
            "K \"Fun Money\"": 200.00,
            "Eating Out": 200.00,
            "Date Budget": 250.00,
            "Travel": 500.00,
            "Home Essentials": 50.00,
            "Gifts": 200.00,
            "Doctors": 100.00,
            "Perscriptions": 50.00,
            "Subscriptions": 100.00,
            "Other": 0.00,
        }
    }

    payload = load_json(path, default_payload)
    if "category_budgets" not in payload or not isinstance(payload["category_budgets"], dict):
        payload = default_payload
    save_json(path, payload)
    budgets = {name: float(value) for name, value in payload["category_budgets"].items()}
    return budgets


def ensure_assignment_store(path: Path) -> dict[str, Any]:
    fallback = {
        "schema_version": 3,
        "assignments": {},
        "last_updated": "",
    }
    payload = load_json(path, fallback)

    assignments: dict[str, dict[str, Any]] = {}

    # Preferred schema: assignments keyed by transaction id.
    for transaction_id, assignment in (payload.get("assignments") or {}).items():
        if not isinstance(transaction_id, str):
            continue
        if isinstance(assignment, dict):
            category = assignment.get("category")
            allocations = assignment.get("allocations")
            assigned_at = assignment.get("assigned_at", "")
        else:
            category = assignment
            allocations = None
            assigned_at = ""

        if isinstance(allocations, list):
            normalized_allocations = []
            for allocation in allocations:
                if not isinstance(allocation, dict):
                    continue
                allocation_category = allocation.get("category")
                allocation_cents = amount_to_cents(allocation.get("amount"))
                if isinstance(allocation_category, str) and allocation_category.strip() and allocation_cents is not None:
                    normalized_allocations.append(
                        {
                            "category": allocation_category.strip(),
                            "amount": cents_to_amount(allocation_cents),
                        }
                    )
            if normalized_allocations:
                assignments[transaction_id] = {
                    "allocations": normalized_allocations,
                    "assigned_at": str(assigned_at or ""),
                }
                continue

        if isinstance(category, str) and category.strip():
            assignments[transaction_id] = {
                "category": category.strip(),
                "assigned_at": str(assigned_at or ""),
            }

    # Legacy schema migration: extract only explicit category decisions.
    for transaction_id, tx in (payload.get("transactions") or {}).items():
        if not isinstance(transaction_id, str) or not isinstance(tx, dict):
            continue
        category = tx.get("category")
        if isinstance(category, str) and category.strip() and transaction_id not in assignments:
            assignments[transaction_id] = {
                "category": category.strip(),
                "assigned_at": "",
            }

    return {
        "schema_version": 3,
        "assignments": assignments,
        "last_updated": payload.get("last_updated", ""),
    }


def collect_csv_files(statements_folder: Path) -> list[Path]:
    if not statements_folder.exists():
        return []
    return sorted([path for path in statements_folder.iterdir() if path.is_file() and path.suffix.lower() == ".csv"])


def collect_transactions_from_csv(
    statements_folder: Path,
    min_date: date = MIN_SUPPORTED_DATE,
) -> dict[str, dict]:
    transaction_map: dict[str, dict] = {}

    for csv_path in collect_csv_files(statements_folder):
        with open(csv_path, "r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                if (row.get("Type") or "").strip().casefold() != "purchase":
                    continue

                date_raw = (row.get("Date") or "").strip()
                if not date_raw:
                    continue

                dt = parse_csv_date(date_raw)
                if dt.date() < min_date:
                    continue

                time_value = normalize_time(row.get("Time") or "")
                if not time_value:
                    continue

                date_value = format_short_date(dt)
                transaction_id = make_transaction_id(date_value, time_value)
                if transaction_id in transaction_map:
                    continue

                merchant = (row.get("Merchant") or "").strip()
                if not merchant:
                    merchant = (row.get("Description") or "").strip()
                if not merchant:
                    continue

                transaction_map[transaction_id] = {
                    "id": transaction_id,
                    "Date": date_value,
                    "Time": time_value,
                    "description": merchant,
                    "amount": parse_amount(row.get("Amount") or ""),
                    "Purchased By": normalize_cardholder(row.get("Cardholder") or ""),
                    "Points": parse_points(row.get("Points") or ""),
                    "source_file": csv_path.name,
                    "year": dt.year,
                    "month": dt.month,
                }
    return transaction_map


def normalize_assignments(assignments: dict[str, dict[str, Any]], valid_categories: list[str]) -> dict[str, dict[str, Any]]:
    valid_set = set(valid_categories)
    cleaned: dict[str, dict[str, Any]] = {}

    for transaction_id, assignment in assignments.items():
        if not isinstance(transaction_id, str) or not isinstance(assignment, dict):
            continue

        allocations = assignment.get("allocations")
        if isinstance(allocations, list):
            cleaned_allocations = []
            for allocation in allocations:
                if not isinstance(allocation, dict):
                    continue
                category = allocation.get("category")
                cents = amount_to_cents(allocation.get("amount"))
                if isinstance(category, str) and category in valid_set and cents is not None:
                    cleaned_allocations.append({"category": category, "amount": cents_to_amount(cents)})
            if cleaned_allocations:
                cleaned[transaction_id] = {
                    "allocations": cleaned_allocations,
                    "assigned_at": assignment.get("assigned_at", ""),
                }
                continue

        category = assignment.get("category")
        if isinstance(category, str) and category in valid_set:
            cleaned[transaction_id] = {
                "category": category,
                "assigned_at": assignment.get("assigned_at", ""),
            }

    return cleaned


def allocations_for_assignment(
    transaction: dict,
    assignment: dict[str, Any],
    valid_categories: set[str],
) -> list[dict[str, Any]] | None:
    transaction_cents = amount_to_cents(transaction.get("amount"))
    if transaction_cents is None or not isinstance(assignment, dict):
        return None

    category = assignment.get("category")
    if isinstance(category, str) and category in valid_categories:
        return [{"category": category, "amount": cents_to_amount(transaction_cents)}]

    allocations = assignment.get("allocations")
    if not isinstance(allocations, list) or len(allocations) < 2:
        return None

    normalized_allocations = []
    allocation_cents_total = 0
    for allocation in allocations:
        if not isinstance(allocation, dict):
            return None
        allocation_category = allocation.get("category")
        allocation_cents = amount_to_cents(allocation.get("amount"))
        if not isinstance(allocation_category, str) or allocation_category not in valid_categories or allocation_cents is None:
            return None
        normalized_allocations.append({"category": allocation_category, "amount": cents_to_amount(allocation_cents)})
        allocation_cents_total += allocation_cents

    return normalized_allocations if allocation_cents_total == transaction_cents else None


def pending_assignment_ids(
    transactions_by_id: dict[str, dict],
    assignments: dict[str, dict[str, Any]],
    categories: list[str],
    only_ids: list[str] | None = None,
) -> list[str]:
    candidates = only_ids if only_ids is not None else list(transactions_by_id.keys())
    valid_set = set(categories)
    pending = []

    for transaction_id in candidates:
        tx = transactions_by_id.get(transaction_id)
        if tx is None:
            continue
        assignment = assignments.get(transaction_id, {})
        if allocations_for_assignment(tx, assignment, valid_set) is None:
            pending.append(transaction_id)

    pending.sort(key=lambda tid: parse_transaction_datetime(transactions_by_id[tid]))
    return pending


@dataclass
class AssignmentUIResult:
    assigned_count: int
    launched: bool
    available: bool


def run_assignment_ui(
    transactions_by_id: dict[str, dict],
    assignments: dict[str, dict[str, Any]],
    category_names: list[str],
    pending_ids: list[str],
) -> AssignmentUIResult:
    if not pending_ids:
        return AssignmentUIResult(assigned_count=0, launched=False, available=True)

    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except Exception:
        return AssignmentUIResult(assigned_count=0, launched=False, available=False)

    tx_map: dict[str, dict] = transactions_by_id
    assigned_count = 0

    root = tk.Tk()
    root.title("Personal Budget Categorizer")
    root.geometry("980x560")

    index = {"value": 0}

    title = tk.Label(root, text="Assign Categories", font=("Helvetica", 18, "bold"))
    title.pack(pady=8)

    progress_var = tk.StringVar(value="")
    detail_var = tk.StringVar(value="")

    progress_label = tk.Label(root, textvariable=progress_var, font=("Helvetica", 12, "bold"))
    progress_label.pack()

    detail_label = tk.Label(root, textvariable=detail_var, justify="left", font=("Helvetica", 12), wraplength=920)
    detail_label.pack(pady=8)

    shortcut_frame = tk.Frame(root)
    shortcut_frame.pack(pady=6)

    button_frame = tk.Frame(root)
    button_frame.pack(fill="both", expand=True, padx=12, pady=8)

    key_symbols = [str(i) for i in range(1, 10)] + [chr(code) for code in range(ord("a"), ord("z") + 1)]
    key_map: dict[str, str] = {}

    def update_view() -> None:
        if index["value"] >= len(pending_ids):
            progress_var.set("All pending new purchases are categorized.")
            detail_var.set("You can close this window.")
            return

        tx = tx_map[pending_ids[index["value"]]]
        progress_var.set(f"{index['value'] + 1} of {len(pending_ids)}")
        detail_var.set(
            f"Date: {tx.get('Date', '')}   Time: {tx.get('Time', '')}   Purchased By: {tx.get('Purchased By', '')}\n"
            f"Merchant: {tx.get('description', '')}\n"
            f"Amount: ${tx.get('amount', 0.0):,.2f}   Points: {tx.get('Points', 0)}"
        )

    def assign_category(category_name: str) -> None:
        nonlocal assigned_count
        if index["value"] >= len(pending_ids):
            return

        tx = tx_map[pending_ids[index["value"]]]
        assignments[tx["id"]] = {
            "category": category_name,
            "assigned_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        }
        assigned_count += 1
        index["value"] += 1
        update_view()

        if index["value"] >= len(pending_ids):
            messagebox.showinfo("Done", "Finished categorizing new purchases.")

    def split_current_transaction() -> None:
        nonlocal assigned_count
        if index["value"] >= len(pending_ids):
            return

        tx = tx_map[pending_ids[index["value"]]]
        transaction_cents = amount_to_cents(tx.get("amount"))
        if transaction_cents is None:
            messagebox.showerror("Cannot split transaction", "The transaction amount is not a valid positive dollar amount.")
            return

        dialog = tk.Toplevel(root)
        dialog.title("Split Transaction")
        dialog.transient(root)
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Split Transaction", font=("Helvetica", 16, "bold")).grid(
            row=0, column=0, columnspan=3, padx=16, pady=(14, 4), sticky="w"
        )
        tk.Label(
            dialog,
            text=f"{tx.get('description', '')} — ${tx.get('amount', 0.0):,.2f}",
            wraplength=560,
            justify="left",
        ).grid(row=1, column=0, columnspan=3, padx=16, pady=(0, 10), sticky="w")

        part_count = tk.IntVar(value=2)
        amount_vars: list[tk.StringVar] = []
        category_vars: list[tk.StringVar] = []
        rows_frame = tk.Frame(dialog)
        rows_frame.grid(row=3, column=0, columnspan=3, padx=16, pady=4, sticky="ew")
        status_var = tk.StringVar(value="")

        def update_split_status(*_args) -> None:
            entered_cents = 0
            all_valid = True
            for amount_var in amount_vars:
                cents = amount_to_cents(amount_var.get().strip())
                if cents is None:
                    all_valid = False
                    continue
                entered_cents += cents

            if not all_valid:
                status_var.set("Enter positive dollar amounts with no fractions of a cent.")
            elif entered_cents == transaction_cents:
                status_var.set("Amounts match the transaction total.")
            else:
                difference = transaction_cents - entered_cents
                direction = "remaining" if difference > 0 else "over"
                status_var.set(f"${format_cents(abs(difference))} {direction}.")

        def rebuild_split_rows() -> None:
            for widget in rows_frame.winfo_children():
                widget.destroy()

            amount_vars.clear()
            category_vars.clear()
            count = max(2, min(10, part_count.get()))
            part_count.set(count)
            equal_cents, remainder = divmod(transaction_cents, count)

            tk.Label(rows_frame, text="Part", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0, 8), pady=3, sticky="w")
            tk.Label(rows_frame, text="Amount", font=("Helvetica", 10, "bold")).grid(row=0, column=1, padx=8, pady=3, sticky="w")
            tk.Label(rows_frame, text="Category", font=("Helvetica", 10, "bold")).grid(row=0, column=2, padx=(8, 0), pady=3, sticky="w")

            for part_index in range(count):
                default_cents = equal_cents + (1 if part_index < remainder else 0)
                amount_var = tk.StringVar(value=format_cents(default_cents))
                category_var = tk.StringVar(value="")
                amount_var.trace_add("write", update_split_status)
                amount_vars.append(amount_var)
                category_vars.append(category_var)

                tk.Label(rows_frame, text=f"{part_index + 1}").grid(row=part_index + 1, column=0, padx=(0, 8), pady=3, sticky="w")
                tk.Entry(rows_frame, textvariable=amount_var, width=12).grid(row=part_index + 1, column=1, padx=8, pady=3, sticky="w")
                ttk.Combobox(
                    rows_frame,
                    textvariable=category_var,
                    values=category_names,
                    state="readonly",
                    width=30,
                ).grid(row=part_index + 1, column=2, padx=(8, 0), pady=3, sticky="w")

            update_split_status()

        tk.Label(dialog, text="Parts:").grid(row=2, column=0, padx=(16, 4), pady=4, sticky="w")
        tk.Spinbox(dialog, from_=2, to=10, textvariable=part_count, width=5, command=rebuild_split_rows).grid(
            row=2, column=1, padx=4, pady=4, sticky="w"
        )
        tk.Button(dialog, text="Apply part count", command=rebuild_split_rows).grid(row=2, column=2, padx=(4, 16), pady=4, sticky="w")
        rebuild_split_rows()

        status_label = tk.Label(dialog, textvariable=status_var, fg="#1F4E78")
        status_label.grid(row=4, column=0, columnspan=3, padx=16, pady=(6, 2), sticky="w")

        def save_split() -> None:
            nonlocal assigned_count
            allocations = []
            allocation_cents_total = 0
            for part_index, (amount_var, category_var) in enumerate(zip(amount_vars, category_vars), start=1):
                cents = amount_to_cents(amount_var.get().strip())
                category = category_var.get().strip()
                if cents is None:
                    messagebox.showerror("Invalid split", f"Part {part_index} needs a positive dollar amount.", parent=dialog)
                    return
                if category not in category_names:
                    messagebox.showerror("Invalid split", f"Choose a category for part {part_index}.", parent=dialog)
                    return
                allocation_cents_total += cents
                allocations.append({"category": category, "amount": cents_to_amount(cents)})

            if allocation_cents_total != transaction_cents:
                difference = transaction_cents - allocation_cents_total
                messagebox.showerror(
                    "Split total does not match",
                    f"Your allocations are ${format_cents(abs(difference))} {'short' if difference > 0 else 'over'} of the purchase total.",
                    parent=dialog,
                )
                return

            assignments[tx["id"]] = {
                "allocations": allocations,
                "assigned_at": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
            }
            assigned_count += 1
            index["value"] += 1
            dialog.destroy()
            update_view()

            if index["value"] >= len(pending_ids):
                messagebox.showinfo("Done", "Finished categorizing new purchases.")

        controls_frame = tk.Frame(dialog)
        controls_frame.grid(row=5, column=0, columnspan=3, padx=16, pady=(6, 14), sticky="ew")
        tk.Button(controls_frame, text="Cancel", command=dialog.destroy).pack(side="left")
        tk.Button(controls_frame, text="Save Split", command=save_split).pack(side="right")

        dialog.wait_window()

    def skip_current() -> None:
        if index["value"] >= len(pending_ids):
            return
        index["value"] += 1
        update_view()

    for widget in shortcut_frame.winfo_children():
        widget.destroy()

    tk.Label(shortcut_frame, text="Keyboard: 1-9 then A-J | Split [S]", font=("Helvetica", 10, "italic")).pack()

    for idx, category_name in enumerate(category_names):
        shortcut = key_symbols[idx] if idx < len(key_symbols) else ""
        if shortcut:
            key_map[shortcut] = category_name

        button_text = f"[{shortcut.upper()}] {category_name}" if shortcut else category_name
        btn = tk.Button(
            button_frame,
            text=button_text,
            width=30,
            anchor="w",
            command=lambda c=category_name: assign_category(c),
        )
        btn.grid(row=idx // 2, column=idx % 2, sticky="ew", padx=6, pady=4)

    button_frame.grid_columnconfigure(0, weight=1)
    button_frame.grid_columnconfigure(1, weight=1)

    control_frame = tk.Frame(root)
    control_frame.pack(fill="x", padx=12, pady=6)

    tk.Button(control_frame, text="Skip [Space]", command=skip_current).pack(side="left")
    tk.Button(control_frame, text="Split [S]", command=split_current_transaction).pack(side="left", padx=8)
    tk.Button(control_frame, text="Save & Close", command=root.destroy).pack(side="right")

    def on_key(event) -> None:
        key = (event.char or "").lower()
        if key in key_map:
            assign_category(key_map[key])
        elif key == "s":
            split_current_transaction()
        elif event.keysym == "space":
            skip_current()

    root.bind("<Key>", on_key)
    update_view()
    root.mainloop()

    return AssignmentUIResult(assigned_count=assigned_count, launched=True, available=True)


def transactions_for_scope(transactions: list[dict], person: str | None = None) -> list[dict]:
    if person is None:
        return list(transactions)
    return [tx for tx in transactions if tx.get("Purchased By") == person]


def compute_budget_rows(transactions: list[dict], category_budgets: dict[str, float], months_count: int) -> tuple[list[dict], dict]:
    rows = []
    total_budget = 0.0

    for category, monthly_budget in category_budgets.items():
        matching_transaction_ids: set[str] = set()
        spent = 0.0
        for tx in transactions:
            for allocation in tx.get("allocations", []):
                if allocation.get("category") == category:
                    matching_transaction_ids.add(tx.get("id", ""))
                    spent += float(allocation.get("amount", 0.0))

        tx_count = len(matching_transaction_ids)
        budget_value = float(monthly_budget) * months_count
        remaining = budget_value - spent
        average = spent / months_count if months_count else spent

        rows.append(
            {
                "category": category,
                "transactions": tx_count,
                "budget": budget_value,
                "spent": spent,
                "remaining": remaining,
                "average": average,
            }
        )

        total_budget += budget_value

    uncategorized = [tx for tx in transactions if not tx.get("allocations")]
    uncat_spent = sum(float(tx.get("amount", 0.0)) for tx in uncategorized)
    uncat_count = len(uncategorized)
    rows.append(
        {
            "category": "Uncategorized",
            "transactions": uncat_count,
            "budget": 0.0,
            "spent": uncat_spent,
            "remaining": -uncat_spent,
            "average": uncat_spent / months_count if months_count else uncat_spent,
        }
    )
    total_transactions = len(transactions)
    total_spent = sum(float(tx.get("amount", 0.0)) for tx in transactions)

    totals = {
        "transactions": total_transactions,
        "budget": total_budget,
        "spent": total_spent,
        "remaining": total_budget - total_spent,
        "average": total_spent / months_count if months_count else total_spent,
    }
    return rows, totals


def write_currency(cell) -> None:
    cell.number_format = "$#,##0.00"


def write_budget_block(worksheet, start_row: int, start_col: int, title: str, rows: list[dict], totals: dict) -> int:
    headers = ["Category", "Transactions", "Budget", "Spent", "Remaining", "Average"]
    end_col = start_col + len(headers) - 1

    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = worksheet.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, color="1F4E78")
    title_cell.fill = PatternFill("solid", fgColor="DCE6F1")

    for i, header in enumerate(headers):
        cell = worksheet.cell(row=start_row + 1, column=start_col + i, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    row_cursor = start_row + 2
    for row in rows:
        worksheet.cell(row=row_cursor, column=start_col, value=row["category"])
        worksheet.cell(row=row_cursor, column=start_col + 1, value=row["transactions"])

        budget_cell = worksheet.cell(row=row_cursor, column=start_col + 2, value=row["budget"])
        spent_cell = worksheet.cell(row=row_cursor, column=start_col + 3, value=row["spent"])
        remain_cell = worksheet.cell(row=row_cursor, column=start_col + 4, value=row["remaining"])
        avg_cell = worksheet.cell(row=row_cursor, column=start_col + 5, value=row["average"])

        for money_cell in (budget_cell, spent_cell, remain_cell, avg_cell):
            write_currency(money_cell)

        if row["remaining"] < 0:
            remain_cell.fill = PatternFill("solid", fgColor="F4CCCC")
        elif row["remaining"] > 0:
            remain_cell.fill = PatternFill("solid", fgColor="D9EAD3")

        row_cursor += 1

    total_label = worksheet.cell(row=row_cursor, column=start_col, value="Total")
    total_label.font = Font(bold=True)
    worksheet.cell(row=row_cursor, column=start_col + 1, value=totals["transactions"]).font = Font(bold=True)

    total_budget_cell = worksheet.cell(row=row_cursor, column=start_col + 2, value=totals["budget"])
    total_spent_cell = worksheet.cell(row=row_cursor, column=start_col + 3, value=totals["spent"])
    total_remaining_cell = worksheet.cell(row=row_cursor, column=start_col + 4, value=totals["remaining"])
    total_avg_cell = worksheet.cell(row=row_cursor, column=start_col + 5, value=totals["average"])

    for money_cell in (total_budget_cell, total_spent_cell, total_remaining_cell, total_avg_cell):
        money_cell.font = Font(bold=True)
        write_currency(money_cell)

    if totals["remaining"] < 0:
        total_remaining_cell.fill = PatternFill("solid", fgColor="F4CCCC")
    elif totals["remaining"] > 0:
        total_remaining_cell.fill = PatternFill("solid", fgColor="D9EAD3")

    return row_cursor + 1


def auto_fit_columns(worksheet, min_width: int = 12, max_width: int = 48) -> None:
    for col_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in col_cells]
        width = max(len(value) for value in values) + 2
        worksheet.column_dimensions[col_cells[0].column_letter].width = max(min_width, min(width, max_width))


def build_personal_budget_workbook(transactions: list[dict], category_budgets: dict[str, float], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook did not create an active worksheet")
    ws.title = "Summary"

    ws["A1"] = "Personal Budget Summary"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = "June 2026 and later | Year sections stack, month sections run left-to-right"
    ws["A2"].font = Font(italic=True, color="666666")

    grouped: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for tx in transactions:
        grouped[(int(tx.get("year", 0)), int(tx.get("month", 0)))].append(tx)

    years = sorted({year for year, _ in grouped.keys()}, reverse=True)
    row = 4

    for year in years:
        year_months = sorted([m for y, m in grouped.keys() if y == year], reverse=True)
        months_count = len(year_months)
        month_sections_start_col = 9
        month_span = 8

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=month_sections_start_col + month_span * max(1, months_count))
        year_cell = ws.cell(row=row, column=1, value=f"YEAR {year}")
        year_cell.font = Font(size=14, bold=True, color="FFFFFF")
        year_cell.fill = PatternFill("solid", fgColor="1F4E78")

        year_transactions = [tx for (y, _), txs in grouped.items() if y == year for tx in txs]

        year_row_cursor = row + 1
        year_rows, year_totals = compute_budget_rows(transactions_for_scope(year_transactions), category_budgets, months_count)
        year_row_cursor = write_budget_block(ws, year_row_cursor, 1, "TOTAL (SAMUEL + KAMRIE)", year_rows, year_totals)

        sam_rows, sam_totals = compute_budget_rows(transactions_for_scope(year_transactions, "Samuel"), category_budgets, months_count)
        year_row_cursor = write_budget_block(ws, year_row_cursor, 1, "SAMUEL", sam_rows, sam_totals)

        kam_rows, kam_totals = compute_budget_rows(transactions_for_scope(year_transactions, "Kamrie"), category_budgets, months_count)
        year_row_cursor = write_budget_block(ws, year_row_cursor, 1, "KAMRIE", kam_rows, kam_totals)

        max_end_row = year_row_cursor

        for month_index, month in enumerate(year_months):
            col = month_sections_start_col + month_index * month_span
            month_transactions = grouped[(year, month)]
            month_label = datetime(year, month, 1).strftime("%B %Y").upper()

            ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 5)
            month_label_cell = ws.cell(row=row + 1, column=col, value=month_label)
            month_label_cell.font = Font(bold=True, color="1F4E78")
            month_label_cell.fill = PatternFill("solid", fgColor="DCE6F1")

            block_cursor = row + 2
            rows_all, totals_all = compute_budget_rows(transactions_for_scope(month_transactions), category_budgets, 1)
            block_cursor = write_budget_block(ws, block_cursor, col, "TOTAL (SAMUEL + KAMRIE)", rows_all, totals_all)

            rows_s, totals_s = compute_budget_rows(transactions_for_scope(month_transactions, "Samuel"), category_budgets, 1)
            block_cursor = write_budget_block(ws, block_cursor, col, "SAMUEL", rows_s, totals_s)

            rows_k, totals_k = compute_budget_rows(transactions_for_scope(month_transactions, "Kamrie"), category_budgets, 1)
            block_cursor = write_budget_block(ws, block_cursor, col, "KAMRIE", rows_k, totals_k)

            if block_cursor > max_end_row:
                max_end_row = block_cursor

        row = max_end_row + 2

    auto_fit_columns(ws)
    wb.save(output_path)


@dataclass
class PersonalBudgetResult:
    transactions_found: int
    pending_before_ui: int
    categorized_in_ui: int
    total_transactions: int
    ui_launched: bool
    ui_available: bool
    rebuild_mode: bool


def transactions_with_assignments(
    transactions_by_id: dict[str, dict],
    assignments: dict[str, dict[str, Any]],
    valid_categories: list[str],
) -> list[dict]:
    valid_set = set(valid_categories)
    output: list[dict] = []

    for transaction_id, tx in transactions_by_id.items():
        tx_copy = dict(tx)
        assignment = assignments.get(transaction_id, {})
        allocations = allocations_for_assignment(tx, assignment, valid_set)
        tx_copy["allocations"] = allocations or []
        output.append(tx_copy)

    return output


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Personal budget categorizer and report exporter")
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="Ignore prior assignment decisions for this run and rebuild assignment decisions from the UI.",
    )
    parser.add_argument(
        "--no-ui",
        action="store_true",
        help="Skip launching Tkinter UI (report will export with current saved decisions only).",
    )
    return parser.parse_args()


def run_personal_budget_pipeline(
    statements_folder: Path,
    categories_file: Path,
    assignments_file: Path,
    output_excel_file: Path,
    launch_ui_for_new: bool = True,
    rebuild: bool = False,
) -> PersonalBudgetResult:
    category_budgets = ensure_categories_file(categories_file)
    category_names = list(category_budgets.keys())
    store = ensure_assignment_store(assignments_file)
    assignments = normalize_assignments(store.get("assignments", {}), category_names)

    if rebuild:
        assignments = {}

    transactions_by_id = collect_transactions_from_csv(statements_folder)
    pending_all_ids = pending_assignment_ids(transactions_by_id, assignments, category_names)

    ui_result = AssignmentUIResult(assigned_count=0, launched=False, available=True)
    if launch_ui_for_new and pending_all_ids:
        ui_result = run_assignment_ui(transactions_by_id, assignments, category_names, pending_all_ids)

    # Persist only assignment decisions (and only after explicit UI interaction or explicit rebuild).
    if ui_result.launched or rebuild:
        store["schema_version"] = 3
        store["assignments"] = assignments
        store["last_updated"] = datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
        save_json(assignments_file, store)

    all_transactions = transactions_with_assignments(transactions_by_id, assignments, category_names)

    build_personal_budget_workbook(all_transactions, category_budgets, output_excel_file)

    return PersonalBudgetResult(
        transactions_found=len(transactions_by_id),
        pending_before_ui=len(pending_all_ids),
        categorized_in_ui=ui_result.assigned_count,
        total_transactions=len(all_transactions),
        ui_launched=ui_result.launched,
        ui_available=ui_result.available,
        rebuild_mode=rebuild,
    )


if __name__ == "__main__":
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    result = run_personal_budget_pipeline(
        statements_folder=base_dir / "statements",
        categories_file=base_dir / "json" / "personal_budget_categories.json",
        assignments_file=base_dir / "json" / "personal_budget_assignments.json",
        output_excel_file=base_dir / "personal_budget_report.xlsx",
        launch_ui_for_new=not args.no_ui,
        rebuild=args.rebuild,
    )
    print(
        "Personal budget export complete | "
        f"transactions_found={result.transactions_found} | "
        f"pending_before_ui={result.pending_before_ui} | "
        f"categorized_in_ui={result.categorized_in_ui} | "
        f"ui_launched={result.ui_launched} | "
        f"ui_available={result.ui_available} | "
        f"total_tracked={result.total_transactions}"
    )
