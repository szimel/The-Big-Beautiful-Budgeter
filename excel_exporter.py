from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from collections import Counter
from collections import defaultdict
import re


SHEET_FORBIDDEN_CHARS = set('[]:*?/\\')
SECTION_COLORS = ["1F4E78", "2E75B6", "4F81BD", "5B9BD5"]
YEAR_PALETTES = [
    {"banner": "1F4E78", "month": "DCE6F1", "card": "FDE9D9"},
    {"banner": "7F6000", "month": "FCE4D6", "card": "FFF2CC"},
    {"banner": "375623", "month": "E2F0D9", "card": "E2EFDA"},
    {"banner": "5B2C6F", "month": "EAD1F2", "card": "EDE2F7"},
    {"banner": "0B6E4F", "month": "D1F2EB", "card": "D6F5E3"},
    {"banner": "9C6500", "month": "FFE9CC", "card": "FFEFD5"},
]


def safe_sheet_title(name: str) -> str:
    cleaned = "".join("_" if char in SHEET_FORBIDDEN_CHARS else char for char in name).strip()
    return cleaned[:31] if cleaned else "Sheet"


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")


def auto_fit_columns(worksheet, min_width: int = 12, max_width: int = 70) -> None:
    for col_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in col_cells]
        width = max(len(value) for value in values) + 2
        worksheet.column_dimensions[col_cells[0].column_letter].width = max(min_width, min(width, max_width))


def set_currency(cell) -> None:
    cell.number_format = "$#,##0.00"


def get_year_palette(year: int, ordered_years: list[int]) -> dict[str, str]:
    index = ordered_years.index(year) if year in ordered_years else 0
    return YEAR_PALETTES[index % len(YEAR_PALETTES)]


def aggregate_year_rows(months: list[dict], person_key: str | None = None) -> tuple[list[dict], dict, dict[str, float]]:
    category_order: list[str] = []
    category_stats: dict[str, dict[str, float]] = {}
    total_transactions = 0
    total_amount = 0.0
    total_points = 0
    month_count = max(1, len(months))

    for month in months:
        if person_key is None:
            rows = month.get("rows", [])
            totals = month.get("totals", {})
        else:
            person_block = month.get("by_person", {}).get(person_key, {})
            rows = person_block.get("rows", [])
            totals = person_block.get("totals", {})

        total_transactions += int(totals.get("transactions", 0) or 0)
        total_amount += float(totals.get("total_amount", 0.0) or 0.0)
        total_points += int(totals.get("points", 0) or 0)

        for item in rows:
            category = item.get("category", "")
            if category not in category_stats:
                category_stats[category] = {"transactions": 0, "total_amount": 0.0}
                category_order.append(category)
            category_stats[category]["transactions"] += int(item.get("transactions", 0) or 0)
            category_stats[category]["total_amount"] += float(item.get("total_amount", 0.0) or 0.0)

    out_rows = []
    avg_map: dict[str, float] = {}
    for category in category_order:
        row_total = category_stats[category]["total_amount"]
        row_avg = row_total / month_count
        avg_map[category] = row_avg
        out_rows.append(
            {
                "category": category,
                "transactions": int(category_stats[category]["transactions"]),
                "total_amount": row_total,
                "average_amount": row_avg,
            }
        )

    totals = {
        "transactions": total_transactions,
        "total_amount": total_amount,
        "points": total_points,
        "average_amount": total_amount / month_count,
    }
    return out_rows, totals, avg_map


def with_yearly_deviation(rows: list[dict], yearly_avg_map: dict[str, float]) -> list[dict]:
    output = []
    for item in rows:
        category = item.get("category", "")
        amount = float(item.get("total_amount", 0.0) or 0.0)
        avg = float(yearly_avg_map.get(category, 0.0))
        output.append(
            {
                "category": category,
                "transactions": int(item.get("transactions", 0) or 0),
                "total_amount": amount,
                "average_amount": avg,
                "deviation": amount - avg,
            }
        )
    return output


def write_summary_table_block(
    worksheet,
    start_row: int,
    start_col: int,
    title: str,
    rows: list[dict],
    palette: dict[str, str],
    thin_border,
    totals: dict | None = None,
    include_deviation: bool = False,
) -> int:
    headers = ["Category", "Transactions", "Total Amount", "Average"]
    if include_deviation:
        headers.append("Deviation")
    end_col = start_col + len(headers) - 1

    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    subheader = worksheet.cell(row=start_row, column=start_col)
    subheader.value = title
    subheader.font = Font(bold=True, color="1F4E78")
    subheader.fill = PatternFill("solid", fgColor=palette["month"])
    subheader.alignment = Alignment(horizontal="left")

    for offset, header in enumerate(headers):
        header_cell = worksheet.cell(row=start_row + 1, column=start_col + offset, value=header)
        apply_header_style(header_cell)

    cursor = start_row + 2
    for item in rows:
        worksheet.cell(row=cursor, column=start_col, value=item.get("category", ""))
        worksheet.cell(row=cursor, column=start_col + 1, value=item.get("transactions", 0))
        amount_cell = worksheet.cell(row=cursor, column=start_col + 2, value=item.get("total_amount", 0.0))
        set_currency(amount_cell)
        avg_cell = worksheet.cell(row=cursor, column=start_col + 3, value=item.get("average_amount", 0.0))
        set_currency(avg_cell)

        if include_deviation:
            deviation = float(item.get("deviation", 0.0) or 0.0)
            deviation_cell = worksheet.cell(row=cursor, column=start_col + 4, value=deviation)
            set_currency(deviation_cell)
            if deviation > 0:
                deviation_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            elif deviation < 0:
                deviation_cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cursor += 1

    if totals is not None:
        total_label = worksheet.cell(row=cursor, column=start_col, value="Total")
        total_label.font = Font(bold=True, color="1F4E78")
        tx_cell = worksheet.cell(row=cursor, column=start_col + 1, value=totals.get("transactions", 0))
        tx_cell.font = Font(bold=True)
        total_amount_cell = worksheet.cell(row=cursor, column=start_col + 2, value=totals.get("total_amount", 0.0))
        total_amount_cell.font = Font(bold=True)
        set_currency(total_amount_cell)
        total_avg_cell = worksheet.cell(row=cursor, column=start_col + 3, value=totals.get("average_amount", 0.0))
        total_avg_cell.font = Font(bold=True)
        set_currency(total_avg_cell)
        if include_deviation:
            worksheet.cell(row=cursor, column=start_col + 4, value="")
        cursor += 1

    for block_row in range(start_row + 1, cursor):
        for block_col in range(start_col, end_col + 1):
            worksheet.cell(row=block_row, column=block_col).border = Border(
                left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
            )

    return cursor


def write_summary_sheet(worksheet, summary_data: list[dict]) -> None:
    worksheet.title = "Summary"
    worksheet["A1"] = "MULTI-YEAR SUMMARY"
    worksheet["A1"].font = Font(size=18, bold=True)
    worksheet["A2"] = "YEARS STACK VERTICALLY. MONTHS SPREAD HORIZONTALLY (NEWEST TO OLDEST)."
    worksheet["A2"].font = Font(color="666666", italic=True)

    row = 4
    thin = Side(style="thin", color="D9D9D9")

    ordered_years = [item.get("year", 0) for item in summary_data]

    for year_block in summary_data:
        year = year_block.get("year", 0)
        palette = get_year_palette(year, ordered_years)
        months = year_block.get("months", [])

        year_rows_all, year_totals_all, year_avg_all = aggregate_year_rows(months, None)
        year_rows_sam, year_totals_sam, year_avg_sam = aggregate_year_rows(months, "Samuel")
        year_rows_kam, year_totals_kam, year_avg_kam = aggregate_year_rows(months, "Kamrie")

        month_table_width = 5
        month_block_span = 6
        months_start_col = 7
        year_table_start_col = 1
        last_month_col = months_start_col + (max(0, len(months) - 1) * month_block_span) + (month_table_width - 1)
        banner_end_col = max(year_table_start_col + 4, last_month_col)

        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=banner_end_col)
        year_cell = worksheet.cell(row=row, column=1)
        year_cell.value = f"YEAR {year}"
        year_cell.font = Font(size=14, bold=True, color="FFFFFF")
        year_cell.fill = PatternFill("solid", fgColor=palette["banner"])
        year_cell.alignment = Alignment(horizontal="left")

        worksheet.merge_cells(start_row=row + 1, start_column=year_table_start_col, end_row=row + 1, end_column=year_table_start_col + 3)
        totals_banner = worksheet.cell(row=row + 1, column=year_table_start_col)
        totals_banner.value = "YEAR TOTALS"
        totals_banner.font = Font(size=12, bold=True, color="FFFFFF")
        totals_banner.fill = PatternFill("solid", fgColor=palette["banner"])
        totals_banner.alignment = Alignment(horizontal="center")

        stats_rows = [
            ("Months", len(months)),
            ("Points", year_totals_all.get("points", 0)),
            ("Samuel Points", year_totals_sam.get("points", 0)),
            ("Kamrie Points", year_totals_kam.get("points", 0)),
        ]
        for offset, (label, value) in enumerate(stats_rows, start=1):
            label_cell = worksheet.cell(row=row + 1 + offset, column=year_table_start_col, value=label)
            label_cell.font = Font(bold=True, color="1F4E78")
            label_cell.fill = PatternFill("solid", fgColor=palette["card"])
            value_cell = worksheet.cell(row=row + 1 + offset, column=year_table_start_col + 1, value=value)
            value_cell.fill = PatternFill("solid", fgColor=palette["card"])

        year_cursor = row + 6
        year_cursor = write_summary_table_block(
            worksheet,
            start_row=year_cursor,
            start_col=year_table_start_col,
            title="TOTAL (SAMUEL + KAMRIE)",
            rows=year_rows_all,
            palette=palette,
            thin_border=thin,
            totals=year_totals_all,
            include_deviation=False,
        )
        year_cursor = write_summary_table_block(
            worksheet,
            start_row=year_cursor,
            start_col=year_table_start_col,
            title="SAMUEL",
            rows=year_rows_sam,
            palette=palette,
            thin_border=thin,
            totals=year_totals_sam,
            include_deviation=False,
        )
        year_cursor = write_summary_table_block(
            worksheet,
            start_row=year_cursor,
            start_col=year_table_start_col,
            title="KAMRIE",
            rows=year_rows_kam,
            palette=palette,
            thin_border=thin,
            totals=year_totals_kam,
            include_deviation=False,
        )

        max_end = year_cursor

        for month_index, month_block in enumerate(months):
            start_col = months_start_col + month_index * month_block_span
            end_col = start_col + month_table_width - 1

            worksheet.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=end_col)
            month_cell = worksheet.cell(row=row + 1, column=start_col)
            month_cell.value = month_block.get("label", "UNKNOWN MONTH")
            month_cell.font = Font(bold=True, color="1F4E78")
            month_cell.fill = PatternFill("solid", fgColor=palette["month"])
            month_cell.alignment = Alignment(horizontal="center")

            all_rows = with_yearly_deviation(month_block.get("rows", []), year_avg_all)
            sam_rows = with_yearly_deviation(month_block.get("by_person", {}).get("Samuel", {}).get("rows", []), year_avg_sam)
            kam_rows = with_yearly_deviation(month_block.get("by_person", {}).get("Kamrie", {}).get("rows", []), year_avg_kam)

            cursor = row + 2
            cursor = write_summary_table_block(
                worksheet,
                start_row=cursor,
                start_col=start_col,
                title="TOTAL (SAMUEL + KAMRIE)",
                rows=all_rows,
                palette=palette,
                thin_border=thin,
                totals=month_block.get("totals", {}),
                include_deviation=True,
            )
            cursor = write_summary_table_block(
                worksheet,
                start_row=cursor,
                start_col=start_col,
                title="SAMUEL",
                rows=sam_rows,
                palette=palette,
                thin_border=thin,
                totals=month_block.get("by_person", {}).get("Samuel", {}).get("totals", {}),
                include_deviation=True,
            )
            cursor = write_summary_table_block(
                worksheet,
                start_row=cursor,
                start_col=start_col,
                title="KAMRIE",
                rows=kam_rows,
                palette=palette,
                thin_border=thin,
                totals=month_block.get("by_person", {}).get("Kamrie", {}).get("totals", {}),
                include_deviation=True,
            )

            if cursor > max_end:
                max_end = cursor

        row = max_end + 2

    auto_fit_columns(worksheet)


def simplify_company_name(description: str) -> str:
    name = description.upper().strip()
    name = re.sub(r"\s+[A-Z]{2}$", "", name)
    name = re.sub(r"#\d+", "", name)
    name = re.sub(r"\b\d{3,}\b", "", name)
    name = re.sub(r"\s+", " ", name).strip(" -*")
    return name or description.upper().strip()


def write_uncategorized_top_companies(worksheet, sections: list[dict]) -> None:
    counter = Counter()
    amount_by_name: dict[str, float] = defaultdict(float)

    for section in sections:
        for tx in section.get("transactions", []):
            name = simplify_company_name(tx.get("description", ""))
            if not name:
                continue
            counter[name] += 1
            amount_by_name[name] += float(tx.get("amount", 0.0))

    top_items = counter.most_common(10)
    start_col = 8

    worksheet.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 2)
    title = worksheet.cell(row=4, column=start_col)
    title.value = "TOP 10 UNCATEGORIZED NAMES"
    title.font = Font(size=12, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center")

    headers = ["Name", "Count", "Total Amount"]
    for offset, label in enumerate(headers):
        header_cell = worksheet.cell(row=5, column=start_col + offset, value=label)
        apply_header_style(header_cell)

    row = 6
    for name, count in top_items:
        worksheet.cell(row=row, column=start_col, value=name)
        worksheet.cell(row=row, column=start_col + 1, value=count)
        amount_cell = worksheet.cell(row=row, column=start_col + 2, value=float(amount_by_name[name]))
        set_currency(amount_cell)
        row += 1


def write_ledger_sheet(worksheet, sheet_name: str, sections: list[dict]) -> None:
    worksheet.title = safe_sheet_title(sheet_name)
    worksheet["A1"] = f"{sheet_name.upper()} LEDGER"
    worksheet["A1"].font = Font(size=18, bold=True)
    worksheet["A2"] = "MONTHLY SECTIONS STACK NEWEST TO OLDEST"
    worksheet["A2"].font = Font(color="666666", italic=True)

    row = 4

    ordered_years = sorted({section.get("year", 0) for section in sections}, reverse=True)

    for index, section in enumerate(sections):
        palette = get_year_palette(section.get("year", 0), ordered_years)
        header_color = palette["banner"]
        label = section.get("label", "UNKNOWN MONTH")
        totals = section.get("totals", {})
        section_transactions = section.get("transactions", [])

        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        header_cell = worksheet.cell(row=row, column=1)
        header_cell.value = label
        header_cell.font = Font(size=13, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", fgColor=header_color)
        header_cell.alignment = Alignment(horizontal="left")

        worksheet.cell(row=row + 1, column=1, value="Transactions")
        worksheet.cell(row=row + 1, column=2, value=totals.get("transactions", 0))
        worksheet.cell(row=row + 1, column=3, value="Total Amount")
        amount_cell = worksheet.cell(row=row + 1, column=4, value=totals.get("total_amount", 0.0))
        set_currency(amount_cell)

        for col in (1, 3):
            metric_cell = worksheet.cell(row=row + 1, column=col)
            metric_cell.font = Font(bold=True, color="1F4E78")

        headers = ["Date", "Description", "Purchased By", "Points", "Amount", "All Matched Categories"]
        for col, text in enumerate(headers, start=1):
            header = worksheet.cell(row=row + 3, column=col, value=text)
            apply_header_style(header)

        data_row = row + 4
        if not section_transactions:
            worksheet.cell(row=data_row, column=2, value="(No transactions)")
            worksheet.cell(row=data_row, column=5, value=0.0)
            set_currency(worksheet.cell(row=data_row, column=5))
            data_row += 1
        else:
            for tx in section_transactions:
                worksheet.cell(row=data_row, column=1, value=tx.get("Date", ""))
                worksheet.cell(row=data_row, column=2, value=tx.get("description", ""))
                worksheet.cell(row=data_row, column=3, value=tx.get("Purchased By", ""))
                worksheet.cell(row=data_row, column=4, value=tx.get("Points", 0))
                amount = worksheet.cell(row=data_row, column=5, value=tx.get("amount", 0.0))
                set_currency(amount)
                worksheet.cell(row=data_row, column=6, value=", ".join(tx.get("categories", [])))
                data_row += 1

        row = data_row + 2

    auto_fit_columns(worksheet)

    if sheet_name.lower() == "uncategorized":
        write_uncategorized_top_companies(worksheet, sections)
        auto_fit_columns(worksheet)


def export_sheet_data_to_excel(
    sheet_data: dict[str, list[dict]], categories_order: list[str], output_file_path: str
) -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    if first_sheet is not None:
        workbook.remove(first_sheet)

    summary_sheet = workbook.create_sheet(title="Summary")
    write_summary_sheet(summary_sheet, sheet_data.get("Summary", []))

    uncategorized_sheet = workbook.create_sheet(title="Uncategorized")
    write_ledger_sheet(uncategorized_sheet, "Uncategorized", sheet_data.get("Uncategorized", []))

    for category in categories_order:
        category_sheet = workbook.create_sheet(title=safe_sheet_title(category))
        write_ledger_sheet(category_sheet, category, sheet_data.get(category, []))

    workbook.save(output_file_path)
